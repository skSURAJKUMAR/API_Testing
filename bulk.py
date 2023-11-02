import openpyxl
import requests
import json
Source_path=r'C:\Users\32020\Downloads\Book45 (1).xlsx'
wb_obj = openpyxl.load_workbook(Source_path)
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row
# Loop will print all columns name
URL=r"http://10.1.8.98:80/PanValidation"
dummy={"Pan_Number": "P","FName": "F","LName": "L"}
count=1
for i in range(2, max_row + 1):
    try:
        Pan_Number=(sheet_obj.cell(row = i, column = 1)).value
        Fname=(sheet_obj.cell(row = i, column = 2)).value
        Lname=(sheet_obj.cell(row = i, column = 3)).value
       
        dummy["Pan_Number"]=Pan_Number
        dummy["FName"]=Fname
        dummy["LName"]=Lname
        dataa= json.dumps(dummy)
        # dummy.replace("P",Pan_Number)
        # dummy.replace("F",Fname)
        # dummy.replace("L",Lname)
 
        print(str(dummy)+"\n")
       
        headers = {
                'Content-Type': 'application/json',
            }
 
        r=requests.post(URL,data=dataa,headers=headers)
        response= r.json()
       
        print("RESPONSE: "+str(response))
    except Exception as e:
        print(e)
        break
